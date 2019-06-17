import openpyxl as xl
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
import logging

import test_Color
import datetime

DebugMessages = False


def print_debug(message):
    if DebugMessages is True:
        print(message)


# - %(levelname)s
logging.basicConfig(filename='app.log', filemode='w', format='%(asctime)s : %(message)s\n',
                    datefmt='%d.%m.%Y, %H:%M:%S',
                    level=logging.DEBUG)

start = ('*' * 20 + ' Start ' + '*' * 20)
end = ('*' * 20 + ' End ' + '*' * 20)

logging.info(start)

def antet_function(name):
    return '*' * 10 + ' ' + name + ' ' + '*' * 10


def get_current_time():
    # return current time
    # %f for ms
    return datetime.datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
    # return datetime.datetime.now().strftime("%d.%m.%Y, %H:%M:%S")


class Prot:
    def __init__(self, filename, sheetname):
        # dictionary used to store all fail test with relevant informations:
        # 'test_name' : {
        # # list of steps with error
        # 'error_line': set(),
        # # total no of errors
        # 'total_errors': None,
        # # total line of test case
        # 'total_lines': None,
        # # total line with errors of test case
        # 'total_lines_error': None,
        # # row of test
        # 'row' : None
        # # test start
        # 'startExecution': None,
        # # test end
        # 'stopExecution': None,
        # # CRs
        # 'CR': [],
        # # comments
        # 'comment': None
        # }
        self.fail_tests = {}

        # file name
        self.file_name = filename

        # sheet name
        self.sheet_name = sheetname

        # connect to workbook
        self.work_book = xl.load_workbook(self.file_name, keep_vba=True, keep_links=True, read_only=False)

        # connect to sheet
        try:
            self.wb_sheet = self.work_book[self.sheet_name]
            self.pass_fail()
        except KeyError:
            print('!!! Sheet {} does not exist in file {}'.format(self.sheet_name, self.file_name))

    def add_fail_test(self, testname):
        logging.info(antet_function('{}({})'.format('add_fail_test', testname)))
        if testname is not None:
            self.fail_tests[testname] = {

                # list of steps with error
                'error_line': set(),

                # total no of errors. One line can have one or more errors
                'total_errors': None,

                # total line of test case
                'total_lines': None,

                # total line with errors of test case
                'total_lines_error': None,

                # row of test
                'row': None,

                # test start
                'startExecution': None,

                # test end
                'stopExecution': None,

                # CRs
                'CR': None,

                # comments
                'comment': None,

                # updated with previous results flag
                'flag': False
            }
            logging.debug('Test {} added to fail_tests dictionary'.format(testname))

    def update_fail_tests(self, testname, member, value):
        # function used to update test informations
        logging.debug(antet_function(
            '{}(testname = {}, member = {}, value = {})'.format('update_fail_tests', testname, member, value)))
        if testname in self.fail_tests:
            logging.debug('Test {} fail, start updates'.format(testname))
            if member == 'error_line':  # add a new element to a set()
                self.fail_tests[testname][member].add(value)
            # elif member == 'CR':  # add a new element to a list()
            #     self.fail_tests[testname][member].append(value)
            else:
                self.fail_tests[testname][member] = value
        logging.debug('Member {} of fail test {} was updated to {}'.format(member, testname, value))

    def read_fail_tests(self, testname):
        # function used to read fail tests:
        #   testname = all -> all tests
        #   testname != all -> specific test will be read
        logging.debug(antet_function('{}({})'.format('read_fail_tests', testname)))
        if testname is not None:
            if testname.lower() == 'all':
                for test_id, test_info in self.fail_tests.items():
                    print("\nTest name: {}".format(test_id))
                    for key in test_info:
                        print('{} : {}'.format(key, test_info[key]))
            elif testname in self.fail_tests:
                print("\nTest name: {}".format(testname))
                for key in self.fail_tests[testname]:
                    print('{} : {}'.format(key, self.fail_tests[testname][key]))
            else:
                print('Test {} not fail'.format(testname))

    def get_sheet_name(self):
        return self.sheet_name

    def get_errors(self, test_name, lineNoSTART_Execution, lineNoSTOP_Execution):
        logging.debug(
            antet_function(
                '{}(test_name = {}, lineNoSTART_Execution = {}, lineNoSTOP_Execution = {})'.format('get_errors',
                                                                                                   test_name,
                                                                                                   lineNoSTART_Execution,
                                                                                                   lineNoSTOP_Execution)))
        # thi function return errors
        sh = self.work_book.worksheets[0]  # first sheet
        errors = 0
        for row in range(lineNoSTART_Execution, lineNoSTOP_Execution + 1):
            if sh.cell(row, 5).value == 'ERROR!!!':
                logging.debug(
                    'test {}, la randul {} am gasit {}'.format(test_name, int(str(sh.cell(row, 1).value).split('!')[1]),
                                                               sh.cell(row, 5).value))
                self.update_fail_tests(testname=test_name, member='error_line',
                                       value=int(str(sh.cell(row, 1).value).split('!')[1]))
                print_debug(
                    'test {}, la randul {} am gasit {}'.format(test_name, int(str(sh.cell(row, 1).value).split('!')[1]),
                                                               sh.cell(row, 5).value))
                errors += 1
        self.update_fail_tests(testname=test_name, member='total_errors', value=errors)
        self.update_fail_tests(testname=test_name, member='total_lines_error',
                               value=len(self.fail_tests[test_name]['error_line']))

    def exist_test(self, test_name):
        logging.debug(antet_function('{}({})'.format('exist_test', test_name)))
        for row in range(3, self.work_book.worksheets[1].max_row + 1):
            if self.work_book.worksheets[1].cell(row, 1).value is not None and test_name in self.work_book.worksheets[
                1].cell(row, 1).value:
                # print_debug("exist_test : test {} exist in {}".format(test_name, self.file_name))
                return True
                break
        else:
            print_debug("exist_test : test {} doesn't exist in {}".format(test_name, self.file_name))

    def test_lines(self, test_name):
        # thi function return test lines
        logging.debug(antet_function('{}({})'.format('test_lines', test_name)))
        if self.exist_test(test_name):
            counter = 4  # max no. of test_name in sheet
            sh = self.work_book.worksheets[0]  # first sheet
            for row in range(1, sh.max_row + 1):
                if test_name in str(sh.cell(row, 1).value):
                    counter -= 1
                    if counter == 2:
                        # test execution start
                        self.update_fail_tests(testname=test_name, member='startExecution', value=row)
                        logging.debug('test execution of test {} started at row {}'.format(test_name, row))
                    elif counter == 0:  # if test_name was found 4th times
                        # test execution finished
                        self.update_fail_tests(testname=test_name, member='stopExecution', value=row)
                        logging.debug('test execution of test {} finished at row {}'.format(test_name, row))
                        for steps in range(row, 0,
                                           -1):  # search for maximum step, from last found test_name to the top in log sheet
                            text_cell = str(sh.cell(steps, 1).value)
                            if 'Step: ' in text_cell:
                                text3 = text_cell.split(':')  # text_cell after split = ['Step', ' 494', '']
                                logging.debug('Test {} has #{} test steps'.format(test_name, int(text3[1])))
                                self.update_fail_tests(testname=test_name, member='total_lines', value=int(text3[1]))
                                break

    def change_sheet(self, sheetname):
        logging.debug(antet_function('{}({})'.format('change_sheet', sheetname)))
        try:
            self.wb_sheet = self.work_book[sheetname]
            return self.wb_sheet
        except KeyError:
            print('Sheet {} does not exist in file {}'.format(sheetname, self.file_name))
            return False

    def pass_fail(self):
        # This function inserts Pass / Fail on column C of sheet
        logging.debug(antet_function('{}({})'.format('pass_fail', self.file_name)))
        for row in range(3, self.work_book.worksheets[1].max_row + 1):
            # 'special' color
            if self.work_book.worksheets[1].cell(row, 2).fill.fgColor.index == 'FFFFC000':
                logging.debug('row {}, test {}, fail'.format(row, self.work_book.worksheets[1].cell(row, 1).value))
                self.work_book.worksheets[1].cell(row, 3).value = 'fail'
                self.add_fail_test(self.work_book.worksheets[1].cell(row, 1).value)
                self.update_fail_tests(self.work_book.worksheets[1].cell(row, 1).value, member='comment',
                                       value=self.work_book.worksheets[1].cell(row, 4).value)
                self.update_fail_tests(self.work_book.worksheets[1].cell(row, 1).value, member='CR',
                                       value=self.work_book.worksheets[1].cell(row, 5).value)
                self.update_fail_tests(self.work_book.worksheets[1].cell(row, 1).value, member='row',
                                       value=row)
                continue
            # if blank cells stop
            elif self.work_book.worksheets[1].cell(row, 2).value is None:
                break
            # insert test status
            elif colors.COLOR_INDEX[
                self.work_book.worksheets[1].cell(row, 2).fill.fgColor.index] in test_Color.test_Pass_Color:
                logging.debug('row {}, test {}, pass'.format(row, self.work_book.worksheets[1].cell(row, 1).value))
                self.work_book.worksheets[1].cell(row, 3).value = 'pass'
            elif colors.COLOR_INDEX[
                self.work_book.worksheets[1].cell(row, 2).fill.fgColor.index] in test_Color.test_Fail_Color:
                logging.debug('row {}, test {}, fail'.format(row, self.work_book.worksheets[1].cell(row, 1).value))
                self.work_book.worksheets[1].cell(row, 3).value = 'fail'
                self.add_fail_test(self.work_book.worksheets[1].cell(row, 1).value)
                self.update_fail_tests(self.work_book.worksheets[1].cell(row, 1).value, member='comment',
                                       value=self.work_book.worksheets[1].cell(row, 4).value)
                self.update_fail_tests(self.work_book.worksheets[1].cell(row, 1).value, member='CR',
                                       value=self.work_book.worksheets[1].cell(row, 5).value)
                self.update_fail_tests(self.work_book.worksheets[1].cell(row, 1).value, member='row',
                                       value=row)
                # # save file
                # tempfile = str(self.file_name).split('.')
                # logging.debug('Saving file {}.{}'.format(tempfile[0] + '_1', tempfile[1]))
                # self.work_book.save(str(tempfile[0] + '_1.' + tempfile[1]))

    def save_file(self):
        # save file
        logging.debug(antet_function('{}({})'.format('save_file', self.file_name)))
        tempfile = str(self.file_name).split('.')
        logging.debug('Saving file {}.{}'.format(tempfile[0] + '_1', tempfile[1]))
        self.work_book.save(str(tempfile[0] + '_' + get_current_time() + '.' + tempfile[1]))
        logging.debug('file {} was saved as {}'.format(self.file_name, str(tempfile[0] + '_' + get_current_time() + '.' + tempfile[1])))


curr_exec = 'prot_1.xlsm'
current_execution = Prot(filename=curr_exec, sheetname='Overview')

prev_exec = 'prot_2.xlsm'
previous_execution = Prot(filename=prev_exec, sheetname='Overview')

# print('Fail tests from {}'.format(current_execution.file_name))
# current_execution.read_fail_tests(testname='all')
# print(current_execution.fail_tests)

# print('Fail tests from {}'.format(previous_execution.file_name))
# previous_execution.read_fail_tests(testname='all')
# print(previous_execution.fail_tests)

for p_id in current_execution.fail_tests:
    print_debug("\nCurrent execution fail test name: {}".format(p_id))

    # check if current fail tests failed also in previous execution
    if p_id in previous_execution.fail_tests.keys():
        logging.info('test {} fail in current execution and also in previous'.format(p_id))

        # check if tests have the same number of lines
        current_execution.test_lines(p_id)
        previous_execution.test_lines(p_id)

        if current_execution.fail_tests[p_id]['total_lines'] == previous_execution.fail_tests[p_id]['total_lines']:
            logging.info('test {} has the same number of lines in current and also in previous execution'.format(p_id))

            # check if tests have the same number of lines with error
            current_execution.get_errors(p_id,
                                         lineNoSTART_Execution=current_execution.fail_tests[p_id]['startExecution'],
                                         lineNoSTOP_Execution=current_execution.fail_tests[p_id]['stopExecution'])
            previous_execution.get_errors(p_id,
                                          lineNoSTART_Execution=previous_execution.fail_tests[p_id]['startExecution'],
                                          lineNoSTOP_Execution=previous_execution.fail_tests[p_id]['stopExecution'])
            if current_execution.fail_tests[p_id]['total_lines_error'] == previous_execution.fail_tests[p_id][
                'total_lines_error']:
                logging.info(
                    'test {} has the same number of lines with error in current and also in previous execution'.format(
                        p_id))
                current_execution.fail_tests[p_id]['comment'] = previous_execution.fail_tests[p_id][
                    'comment']
                current_execution.fail_tests[p_id]['CR'] = previous_execution.fail_tests[p_id][
                    'CR']
                current_execution.fail_tests[p_id]['flag'] = True
            else:
                logging.debug(
                    'test {} NOT the same number of lines with error in current and in previous execution'.format(
                        p_id))
        else:
            logging.info('test {} NOT the same number of lines in current and in previous execution'.format(p_id))
    else:
        logging.info('test {} fail only in current execution'.format(p_id))

for test_name in current_execution.fail_tests:
    # print('Test name: ', test_name)
    if current_execution.fail_tests[test_name]['flag']:
        row = current_execution.fail_tests[test_name]['row']
        current_execution.work_book.worksheets[1].cell(row, 4).value = str(
            current_execution.fail_tests[test_name]['comment'])
        current_execution.work_book.worksheets[1].cell(row, 5).value = current_execution.fail_tests[test_name]['CR']

current_execution.save_file()
previous_execution.save_file()

logging.info(end)